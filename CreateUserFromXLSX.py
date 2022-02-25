import openpyxl as xl
import requests
import base64

url = 'https://iljira.icicilombard.com/rest/api/2/user'

sample_string = "jira-admin:admin"
sample_string_bytes = sample_string.encode("ascii")
base64_bytes = base64.b64encode(sample_string_bytes)
base64_string = base64_bytes.decode("ascii")

headers = {
    "Content-Type": "application/json",
    "Authorization": f"Basic {base64_string}"
}

wb = None
try:
    wb = xl.open('jiju.xlsx')

    sheet = wb['Sheet1']
    i = 0
    for row in range(2,sheet.max_row+1):
        fullname = sheet.cell(row,1)
        username = sheet.cell(row,2)
        email = sheet.cell(row,3)
        password = sheet.cell(row, 4)
        application = sheet.cell(row, 5)

        create_user = {
            "name": username.value,
            "password": password.value,
            "emailAddress": email.value,
            "displayName": fullname.value,
            "applicationKeys": [
                application.value
            ]
        }
        res = requests.post(url=url, json=create_user, headers=headers)

        print(res.status_code)
        print(res.json())
        i += 1
    print(i,'users created')

        # print(fullname.value, username.value, email.value, password.value, application.value)
except Exception as e:
    print(e)
finally:
    if wb is not None:
        wb.close()